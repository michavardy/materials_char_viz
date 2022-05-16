import pandas as pd
from pathlib import Path
import numpy as np
import re
import math
import itertools
import openpyxl
from openpyxl.styles.alignment import Alignment
import matplotlib.pyplot as plt
import seaborn as sns

"""
TODO
    1. add ceramics
"""

class Material_Database():
    """
        mat.materials - is dictionary with all of the material dataframes
        mat.materials_extract_keys - is list of keys to materials dictionary used for extraction

        mat.keyword_dataframes_dict - is dictionary of dataframes extracted by keyword
        mat.extract_index_map - is dictionary mapping extract_index to ls_dyna material type
    """
    def __init__(self, materials, materials_extract_keys, add_erosion, keywords):
        self.materials = materials
        self.materials_extract_keys = materials_extract_keys
        self.add_erosion = add_erosion
        self.keywords = keywords
        self.keyword_dataframes_dict = {}
        self.extract_index = 1
        self.extract_index_map ={}
        if type(self.materials)==dict:
            self.format_materials()
        self.iterate_through_extract_keys()
    def format_materials(self):
        material_library_name = self.materials.keys()
        materials = {}
        for name in material_library_name:
            for prop in self.materials_extract_keys + self.add_erosion:
                df = self.materials[name][prop]
                df.db = name
                if 'title' in df.columns:
                    df.title = df.title.apply(lambda t:f'{t}_{name}')
                df.mid = df.mid.apply(lambda m:f'{m}_{name}')
                if prop in materials.keys():
                    materials[prop].append(df)
                else:
                    materials[prop] = [df]
        self.compound_materials = materials
        self.merge_compound_materials()
    def drop_unnamed(self):
        new_list=[]
        for df in self.df_list:
            if 'Unnamed: 0' in df.columns:
                df = df.drop(['Unnamed: 0'], axis=1)
            new_list.append(df)
        self.df_list = new_list
    def merge_compound_materials(self):
        materials = {}
        for key in self.compound_materials:
            self.df_list = self.compound_materials[key]
            self.drop_unnamed()
            materials[key] = pd.concat(self.df_list)
        self.materials = materials
    def iterate_through_extract_keys(self):
        for extract in self.materials_extract_keys:
            self.extract = extract
            self.gen_mat_df()
            self.gen_add_erosion_df()
            self.mat_df = self.mat_df.merge(self.add_erosion_df, how='left', on='mid')
            self.material_keyword_dicts = {}          
            self.extract_keywords()
            self.build_dataframes_from_keywords()
            self.extract_index = self.extract_index + 1
    def gen_mat_df(self):
        self.mat_df = self.materials[self.extract]
        if 'Unnamed: 0' in self.mat_df.columns:
            self.mat_df = self.mat_df.drop(['Unnamed: 0'],axis=1)
        self.mat_df = self.mat_df.drop_duplicates()
    def extract_keywords(self):
        """
            this function builds a dictionary:
                keys: material_model, material_name, dict
                values: dictionaries of mid: material_name
            this dictionary called material_keyword_dicts can be used
            to build dataframes around given materials
        """
        for k in self.keywords:
            if type(k) == list:
                for i in k:
                    if f'{k[0]}_dict' in  self.material_keyword_dicts:
                        self.material_keyword_dicts[f'{k[0]}_dict'] = \
                            {**self.material_keyword_dicts[f'{k[0]}_dict'], 
                            **self.mat_df.title[self.mat_df.title.str.contains(i,flags=re.IGNORECASE)].to_dict()}
                    else:
                        self.material_keyword_dicts[f'{k[0]}_dict'] =  \
                            self.mat_df.title[self.mat_df.title.str.contains(i,flags=re.IGNORECASE)].to_dict()
            else:
                if f'{k}_dict' in  self.material_keyword_dicts:
                    self.material_keyword_dicts[f'{k}_dict'] = \
                        {**self.material_keyword_dicts[f'{k}_dict'],
                         **self.mat_df[self.mat_df.str.contains(i,flags=re.IGNORECASE)].to_dict()}
                else:
                    self.material_keyword_dicts[f'{k}_dict'] = \
                        self.mat_df.title[self.mat_df.title.str.contains(k,flags=re.IGNORECASE)].to_dict()
    def extract_mid_dict_df(self):
        self.mid_list = list(self.mid_dict.keys())
        self.mid_dict_df = self.mat_df.loc[self.mid_list]
    def build_dataframes_from_keywords(self):
        self.keyword_keys = [f'{i}_dict' if type(i)==str else f'{i[0]}_dict' for i in self.keywords]
        for key in self.keyword_keys:
            self.mid_dict = self.material_keyword_dicts[key]
            self.extract_mid_dict_df()
            self.mid_dict_df['isIdentical']=[[] for i in range(len(self.mid_dict_df))]
            self.check_identical()
            self.hide_identical()
            self.keyword_dataframes_dict[f'{key}_{self.extract_index}']=self.mid_dict_df
            self.extract_index_map[self.extract_index] = self.extract
    def extract_add_erosion_ser(self):
        add_erosion_ser = self.ae_df[self.ae_df.mid == self.mid].iloc[0]
        add_erosion_ser = add_erosion_ser.dropna()
        add_erosion_ser = add_erosion_ser[add_erosion_ser!=0]
        if 'title' in list(add_erosion_ser):
            add_erosion_ser = add_erosion_ser.drop(['title'])
        if 'Unnamed: 0' in add_erosion_ser:
            add_erosion_ser = add_erosion_ser.drop(['Unnamed: 0'])
        return(add_erosion_ser)
    def gen_add_erosion_df(self):
        #iterate through add erosion dfs
        self.add_erosion_df = pd.DataFrame()
        for mid in self.mat_df.mid:
            for ae in self.add_erosion:
                self.mid = mid
                self.ae_df = self.materials[ae]
                if self.mid in list(self.ae_df.mid):
                    #print(add_erosion_ser)
                    add_erosion_ser = self.extract_add_erosion_ser()
                    self.add_erosion_ser = add_erosion_ser
                    self.add_erosion_df = \
                        self.add_erosion_df.append(add_erosion_ser)
        if "title" in self.add_erosion_df.columns:
            self.add_erosion_df = self.add_erosion_df.drop(['title'], axis=1)
    def check_identical(self):
        self.combo = list(itertools.combinations(range(len(self.mid_dict_df)),2))
        for c in self.combo:
            df = self.mid_dict_df.drop(['mid','title'], axis=1)
            if all(df.iloc[c[0]].fillna(0) == df.iloc[c[1]].fillna(0)):

                new_list0 = [i for i in self.mid_dict_df.iat[c[0], -1]]
                new_list0.append(self.mid_dict_df.iloc[c[1]].mid)
                self.mid_dict_df.iat[c[0], -1] = new_list0

                new_list1 = [i for i in self.mid_dict_df.isIdentical.iloc[c[1]]]
                new_list1.append(self.mid_dict_df.iloc[c[0]].mid)
                self.mid_dict_df.iat[c[1], -1] =new_list1

                    
                #print(self.mid_dict_df.isIdentical, self.mid_dict_df.iloc[c[0]].mid)
    def hide_identical(self):
        skip_list = []
        for index, row in self.mid_dict_df.iterrows():
            if row.isIdentical:
                if row.mid in skip_list:
                    continue
                else:
                    skip_list = skip_list + row.isIdentical
        self.mid_dict_df = self.mid_dict_df[~self.mid_dict_df.mid.isin(skip_list)]
    def dump(self,path):
        writer = pd.ExcelWriter(path, engine='xlsxwriter')
        for key in self.keyword_dataframes_dict.keys():
            if len(self.keyword_dataframes_dict[key])>1:
                df = self.keyword_dataframes_dict[key].set_index('title').T
                df.to_excel(writer, sheet_name=key)
        writer.save()
class Database_Format():
    def __init__(self,xl_dir):
        self.xl_dir = xl_dir
        self.wb = openpyxl.load_workbook(self.xl_dir)
        self.define_styles()
        self.iterate_through_worksheets()
        self.wb.save(f'{Path(self.xl_dir).stem}_formated.xlsx')
    def define_styles(self):
        self.my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
        self.grey = openpyxl.styles.colors.Color(rgb='00B2BABB')
        self.bgrey = openpyxl.styles.colors.Color(rgb='009FCACF')
        self.my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=self.my_red)
        self.header_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=self.grey)
        self.props_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=self.bgrey)
        self.border = openpyxl.styles.borders.Border(outline=True)
    def iterate_through_worksheets(self):
        for ws in self.wb.worksheets:
            self.ws = ws
            self.find_plot_path()
            self.rows = [row for row in self.ws.rows]
            self.global_format()
            self.format_headers()
            self.format_props()
            self.iterate_over_sheet_rows()
            if self.plot_path:
                self.img = openpyxl.drawing.image.Image(self.plot_path)
                self.img.anchor = 'A30'
                self.ws.add_image(self.img)
    def format_headers(self):
        self.header_row = self.rows[0]
        #self.set_column_width()
        for cell in self.header_row[1:]:
            cell.alignment = Alignment(wrap_text=True,horizontal='center', vertical='center' )
            self.ws.column_dimensions[cell.column_letter].width = 20
            cell.fill = self.header_fill
            #cell.style=self.border
    def format_props(self):
        self.props_column = [i for i in self.ws.columns][0]
        for cell in self.props_column:
            cell.fill = self.props_fill
    def center_all_content(self):
        #center all content
        for row in self.rows:
            for cell in row:
                cell.alignment = Alignment(horizontal='center')
                #cell.style = self.border
    def global_format(self):
        self.center_all_content()
    def iterate_over_sheet_rows(self):
        self.rows = [row for index,row in enumerate(self.ws.rows) \
             if index not in [0,1]]
        for row in self.rows:
            # skip over row headers
            self.row = row[1:]
            self.format_row()
    def find_plot_path(self):
        if self.ws.title in [i.stem for i in list(Path('C:\\Users\\Micha.Vardy\\Desktop\\pptx\\figs').iterdir())]:
            self.plot_path = r'C:\Users\Micha.Vardy\Desktop\pptx\figs\{title}.png'.format(title=self.ws.title)
        else:
            self.plot_path = False
    def format_row(self):
        self.values = pd.Series([v.value for v in self.row])
        duplicate_mask = self.values.duplicated()
        unique_indexes = self.values[~duplicate_mask].index.to_list()
        unique_cells = [v for i,v in enumerate(self.row) if i in unique_indexes][1:]
        for cell in unique_cells:
            cell.fill = self.my_fill
class Plot_jc():
    def __init__(self,xl_dir):
        self.xl_dir = xl_dir
        self.df_dict =  pd.read_excel(self.xl_dir,sheet_name=None)
        self.df_dict = {key:{'df':self.df_dict[key]} for key in self.df_dict.keys()}
        self.check_material_type()
        self.plot_jc()
    def check_material_type(self):
        for key in self.df_dict.keys():
            self.df_dict[key]['isJC'] = self.check_jc(self.df_dict[key]['df'])
            self.df_dict[key]['isModified'] = self.check_modified_jc(self.df_dict[key]['df']) 
            self.df_dict[key]['isJH'] = self.check_jh(self.df_dict[key]['df']) 
    def check_jc(self,df):
        self.jc_check = pd.DataFrame({
            'a': [any(df['Unnamed: 0'].str.contains('^a$'))],
            'b': [any(df['Unnamed: 0'].str.contains('^b$'))],
            'c': [any(df['Unnamed: 0'].str.contains('^c$'))],
            'n': [any(df['Unnamed: 0'].str.contains('^n$'))],
            'm': [any(df['Unnamed: 0'].str.contains('^m$'))],
        })
        self.jc_damage_check = pd.DataFrame({
            'd1': [any(df['Unnamed: 0'].str.contains('^d1$'))],
            'd2': [any(df['Unnamed: 0'].str.contains('^d2$'))],
            'd3': [any(df['Unnamed: 0'].str.contains('^d3$'))],
            'd4': [any(df['Unnamed: 0'].str.contains('^d4$'))]
        })
        return(all(pd.concat([self.jc_check.iloc[0],self.jc_damage_check.iloc[0]])))
    def check_modified_jc(self,df):
        self.modified_jc = pd.DataFrame({
            'a/siga': [any(df['Unnamed: 0'].str.contains('^a/siga$'))],
            'b/b': [any(df['Unnamed: 0'].str.contains('^b/b$'))],
            'c/beta1': [any(df['Unnamed: 0'].str.contains('^c/beta1$'))],
            'n/beta0': [any(df['Unnamed: 0'].str.contains('^n/beta0$'))],
            'm/na': [any(df['Unnamed: 0'].str.contains('^m/na$'))],
        })
        return(all(self.modified_jc.iloc[0]))
    def check_jh(self,df):
        self.jh_check = pd.DataFrame({
            'a': [any(df['Unnamed: 0'].str.contains('^a$'))],
            'b': [any(df['Unnamed: 0'].str.contains('^b$'))],
            'c': [any(df['Unnamed: 0'].str.contains('^c$'))],
            'n': [any(df['Unnamed: 0'].str.contains('^n$'))],
            'm': [any(df['Unnamed: 0'].str.contains('^m$'))],
        })
        return(all(self.jh_check.iloc[0]))
    def plot_jc_axis_pre_formatting(self):
        self.figjc, (self.ax1, self.ax2) = plt.subplots(1, 2,figsize=(20,10))
        self.figjc.suptitle(f'{re.split("_",self.mat_type)[0]}_jc_plots', fontsize=16)
        self.ax1.set_title('plastic stress strain')
        self.ax1.set_xlabel('strain')
        self.ax1.set_ylabel('stress [Mpa]')
        self.ax2.set_title('damage')
        self.ax2.set_xlabel('triax')
        self.ax2.set_ylabel('failure strain')
        self.ax2.invert_xaxis()
    def plot_modified_axis_pre_formatting(self):
        self.fig, self.ax1 = plt.subplots(1,1,figsize=(20,10))
        self.fig.suptitle(f'{re.split("_",self.mat_type)[0]}_jc_plots', fontsize=16)
        self.ax1.set_title('plastic stress strain')
        self.ax1.set_xlabel('strain')
        self.ax1.set_ylabel('stress [Mpa]')
    def group_duplicates(self,df,row_names):
        #self.groups = [", ".join(group.index.to_list()) for val,group in df.T.groupby(row_names)]
        self.groups = [group.index.to_list() for val,group in df.T.groupby(row_names)]
        self.vals = [val for val,group in df.T.groupby(row_names)]
        return()
    def plot_jc_mat(self,**kwargs):
        #kwargs['plastic']
        self.plot_jc_axis_pre_formatting()
        self.df = self.df_dict[self.mat_type]['df']
        self.plastic_kwargs = kwargs['plastic']
        self.df_plastic = self.df.set_index(['Unnamed: 0']).loc[kwargs['plastic']]
        self.group_duplicates(self.df_plastic,kwargs['plastic'].to_list())
        self.df_plastic = self.df_plastic[[i[0] for i in self.groups]]
        self.df_damage = self.df.set_index(['Unnamed: 0']).loc[kwargs['damage']]
        self.f = pd.DataFrame({"strain":[i/100 for i in range(10)] + [i/10 for i in range(1,31)]})
        self.d = pd.DataFrame({"triax":[-i/10 for i in range(20,0,-1)] + [i/10 for i in range(0,21)]})
        for  index,col in enumerate(self.df_plastic.columns):
            markers=[',', '+', '.', 'o', '*',"v","^","<",">","1",'2','3','4','8','s','p','P','h','H']
            df = self.df_plastic[col]
            self.f[col] = df.a + df.b*self.f.strain**df.n
            self.ax1.plot(self.f.strain,self.f[col], label=col, marker=markers[index])
            self.ax1.legend()
            self.figjc.text(0.5, 0.5,"\n\n".join(["\n".join(i) for i in self.groups]),
                verticalalignment='bottom', 
                horizontalalignment='center',
                fontsize=15 )
            self.df1 = self.df_damage[col]
            self.d[col] = self.df1.d1 + self.df1.d2*np.exp(self.df1.d3*self.d.triax)
            self.ax2.plot(self.d.triax,self.d[col], label=col)
            self.ax2.legend()
            self.figjc.savefig(f'C:\\Users\\Micha.Vardy\\Desktop\\pptx\\figs\\{self.mat_type}.png')
            self.figjc.savefig(f'C:\\Users\\Micha.Vardy\\Desktop\\pptx\\figs\\{self.mat_type}.svg')
    def plot_jh_mat(self,**kwargs):
        self.plot_modified_axis_pre_formatting()
        self.df = self.df_dict[self.mat_type]['df']
        self.plastic_kwargs = kwargs['plastic']
        self.df_plastic = self.df.set_index(['Unnamed: 0']).loc[kwargs['plastic']]
        self.group_duplicates(self.df_plastic,kwargs['plastic'].to_list())
        self.df_plastic = self.df_plastic[[i[0] for i in self.groups]]
        self.f = pd.DataFrame({"strain":[i/100 for i in range(10)] + [i/10 for i in range(1,31)]})
        for  index,col in enumerate(self.df_plastic.columns):
            markers=[',', '+', '.', 'o', '*',"v","^","<",">","1",'2','3','4','8','s','p','P','h','H']
            df = self.df_plastic[col]
            self.f[col] = df.a + df.b*self.f.strain**df.n
            self.ax1.plot(self.f.strain,self.f[col], label=col, marker=markers[index])
            self.ax1.legend()
            self.fig.text(0.5, 0.1,"\n\n".join(["\n".join(i) for i in self.groups]),
                verticalalignment='bottom', 
                horizontalalignment='center',
                fontsize=15 )
            self.fig.savefig(f'C:\\Users\\Micha.Vardy\\Desktop\\pptx\\figs\\{self.mat_type}.png')
            self.fig.savefig(f'C:\\Users\\Micha.Vardy\\Desktop\\pptx\\figs\\{self.mat_type}.svg')
    def plot_modified_jc_mat(self,**kwargs):
        self.plot_modified_axis_pre_formatting()
        self.df = self.df_dict[self.mat_type]['df']
        self.plastic_kwargs = kwargs['plastic']
        self.df_plastic = self.df.set_index(['Unnamed: 0']).loc[kwargs['plastic']]
        self.group_duplicates(self.df_plastic,kwargs['plastic'].to_list())
        self.df_plastic = self.df_plastic[[i[0] for i in self.groups]]
        self.f = pd.DataFrame({"strain":[i/100 for i in range(10)] + [i/10 for i in range(1,31)]})
        for  index,col in enumerate(self.df_plastic.columns):
            markers=[',', '+', '.', 'o', '*',"v","^","<",">","1",'2','3','4','8','s','p','P','h','H']
            df = self.df_plastic[col]
            self.f[col] = df['a/siga'] + df['b/b']*self.f.strain**df['n/beta0']
            self.ax1.plot(self.f.strain,self.f[col], label=col, marker=markers[index])
            self.ax1.legend()
            self.fig.text(0.5, 0.5,"\n\n".join(["\n".join(i) for i in self.groups]),
                verticalalignment='bottom', 
                horizontalalignment='center',
                fontsize=15 )
            self.fig.savefig(f'C:\\Users\\Micha.Vardy\\Desktop\\pptx\\figs\\{self.mat_type}.png')
            self.fig.savefig(f'C:\\Users\\Micha.Vardy\\Desktop\\pptx\\figs\\{self.mat_type}.svg')
    def plot_jc(self):
        self.df_dict_jc = {key:self.df_dict[key]['df'] for key in self.df_dict if self.df_dict[key]['isJC']}
        self.df_dict_modified_jc = {key:self.df_dict[key]['df'] for key in self.df_dict if self.df_dict[key]['isModified']}
        self.df_dict_jh = {key:self.df_dict[key]['df'] for key in self.df_dict if self.df_dict[key]['isJH']}
        #self.df_dict_jc_test = [list(self.df_dict_jc.keys())[0]]
        for mat_type in self.df_dict_modified_jc.keys():
            self.mat_type = mat_type
            print(self.mat_type)
            self.plot_modified_jc_mat(**{'plastic' : self.modified_jc.keys()})
        for mat_type in self.df_dict_jh.keys():
            self.mat_type = mat_type
            print(self.mat_type)
            self.plot_jh_mat(**{'plastic' : self.jh_check.keys()})
        for mat_type in self.df_dict_jc.keys():
            self.mat_type = mat_type
            print(self.mat_type)
            self.plot_jc_mat(**{
                'plastic' : self.jc_check.keys(), 
                'damage' : self.jc_damage_check.keys()})
if __name__ =="__main__":
    mdb = {}
    mdb['materials'] = {
            'zvika':pd.read_excel(r"C:\Users\Micha.Vardy\Desktop\pptx\zvika_materials.xlsx",sheet_name=None),
            'micha':pd.read_excel(r"C:\Users\Micha.Vardy\Desktop\pptx\balistic_mat_v2.xlsx",sheet_name=None),
            'FV':pd.read_excel(r"C:\Users\Micha.Vardy\Desktop\pptx\FV_materials.xlsx",sheet_name=None),
            'gabbi':pd.read_excel(r"C:\Users\Micha.Vardy\Desktop\pptx\Gabi_materials.xlsx",sheet_name=None),
            #'cassets':pd.read_excel(r"C:\Users\Micha.Vardy\Desktop\pptx\balistic_mat_cassets.xlsx",sheet_name=None)
            }       
    mdb['materials_extract_keys'] = ['MAT_JOHNSON_COOK_TITLE', 'MAT_MODIFIED_JOHNSON_COOK_TITL','MAT_JOHNSON_HOLMQUIST_CERAMICS']
    mdb['add_erosion'] = ['MAT_ADD_EROSION_TITLE', 'MAT_ADD_EROSION']
    mdb['keywords'] = [['Ti']]
    #mat = Material_Database(**mdb)
    #mat.dump('materials_output.xlsx')
    Path(f"C:\\Users\\Micha.Vardy\\Desktop\\pptx\\figs")
    plot = Plot_jc('materials_output.xlsx')
    form = Database_Format('materials_output.xlsx')

    #mdb['materials']['cassets'] = {i: mdb['materials']['cassets'][i] for i in list(mdb['materials']['cassets'].keys()) if 'mid' in mdb['materials']['cassets'][i].columns}
    #mdb['materials_extract_keys'] =  [
    #                                    'MAT_CRUSHABLE_FOAM_TITLE',
    #                                    'MAT_JOHNSON_COOK_TITLE',
    #                                    'MAT_NULL_TITLE',
    #                                    'MAT_MODIFIED_JOHNSON_COOK_TITL',
    #                                    'MAT_HIGH_EXPLOSIVE_BURN_TITLE',
    #                                    'MAT_LOW_DENSITY_FOAM_TITLE',
    #                                    'MAT_HONEYCOMB_TITLE',
    #                                    'MAT_JOHNSON_HOLMQUIST_CERAMICS',
    #                                    'MAT_COMPOSITE_FAILURE_SOLID_MO',
    #                                    'MAT_JOHNSON_HOLMQUIST_JH1_TITL',
    #                                    'MAT_SIMPLIFIED_JOHNSON_COOK_TI',
    #                                    'MAT_POWER_LAW_PLASTICITY_TITLE',
    #                                    'MAT_PLASTIC_KINEMATIC_TITLE']
